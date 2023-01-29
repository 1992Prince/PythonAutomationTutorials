import os.path

import openpyxl
import pytest
from selenium import webdriver


@pytest.fixture(scope="function")
def chrome_browser():
    global driver
    driver = webdriver.Chrome(executable_path="../driver/chromedriver.exe")
    driver.get("https://facebook.com")
    driver.maximize_window()

    yield driver
    driver.quit()

@pytest.mark.hookwrapper
def pytest_runtest_makereport(item):
    """
        Extends the PyTest Plugin to take and embed screenshot in html report, whenever test fails.
        :param item:
        """
    pytest_html = item.config.pluginmanager.getplugin('html')
    outcome = yield
    report = outcome.get_result()
    extra = getattr(report, 'extra', [])

    if report.when == 'call' or report.when == "setup":
        extra.append(pytest_html.extras.url("http://www.vitria.com/"))
        xfail = hasattr(report, 'wasxfail')
        if (report.skipped and xfail) or (report.failed and not xfail):
            file_name = report.nodeid.replace("::", "_") + ".png"
            _capture_screenshot(file_name)
            if file_name:
                html = '<div><img src="%s" alt="screenshot" style="width:304px;height:228px;" ' \
                       'onclick="window.open(this.src)" align="right"/></div>' % file_name
                extra.append(pytest_html.extras.html(html))
        report.extra = extra


def _capture_screenshot(name):
        driver.get_screenshot_as_file(name)

def pytest_html_report_title(report):
    report.title = "OI-Vitria Automation Report"

def get_data(testcasename):
    # return [
    #     ["pkp@mail.com","pkp123"],
    #     ["rkp@mail.com","mkp123"],
    #     ["skp@mail.com","skp123"]
    # ]
    workbook = openpyxl.load_workbook("..//testData/PythonTestData.xlsx")
    sheet = workbook["Sheet1"]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    print("rows :",totalrows,"---cols :",totalcols)
    mainList = []

    for i in range(2,totalrows+1):
        dataList = []
        if sheet.cell(row=i,column=1).value == testcasename:
            for j in range(2,totalcols+1):
                data = sheet.cell(row=i,column=j).value
                dataList.insert(j,data)
            mainList.insert(i,dataList)
    print("Main List ",mainList)
    return mainList
