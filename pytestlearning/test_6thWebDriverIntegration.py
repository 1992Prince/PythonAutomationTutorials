import time

import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By

from pytestlearning.conftest import get_data


# @pytest.fixture()
# def log_on_failure(request):
#
#     yield
#     item = request.node
#     if item.rep_call.failed:
#         pass



# def get_data(testcasename):
#     # return [
#     #     ["pkp@mail.com","pkp123"],
#     #     ["rkp@mail.com","mkp123"],
#     #     ["skp@mail.com","skp123"]
#     # ]
#     workbook = openpyxl.load_workbook("..//testData/PythonTestData.xlsx")
#     sheet = workbook["Sheet1"]
#     totalrows = sheet.max_row
#     totalcols = sheet.max_column
#     print("rows :",totalrows,"---cols :",totalcols)
#     mainList = []
#
#     for i in range(2,totalrows+1):
#         dataList = []
#         if sheet.cell(row=i,column=1).value == testcasename:
#             for j in range(2,totalcols+1):
#                 data = sheet.cell(row=i,column=j).value
#                 dataList.insert(j,data)
#             mainList.insert(i,dataList)
#     print("Main List ",mainList)
#     return mainList



@pytest.mark.parametrize("username,password",get_data("Test1"))
def test_doLogin(username,password,chrome_browser):
     driver = chrome_browser
     driver.find_element(By.ID,"email").send_keys(username)
     driver.find_element(By.ID, "pass").send_keys(password)
    #driver.find_element(By.XPATH,"id").click()
     time.sleep(2)


