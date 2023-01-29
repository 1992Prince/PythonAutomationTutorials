import openpyxl

def getRowCount(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColCount(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column

def getCellData(path,sheetName,rownum,colnum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=colnum).value


def setCellData(path,sheetName,rownum,colnum,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(path)


