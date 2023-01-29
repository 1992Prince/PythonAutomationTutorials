import openpyxl

workbook = openpyxl.load_workbook("..//testData/PythonTestData.xlsx")
sheet = workbook["Sheet1"]
totalrows = sheet.max_row
totalcols = sheet.max_column

print("Total rows are :",totalrows," total columns are :",totalcols)

print(sheet.cell(row=2,column=1).value)

print("Printing complete excel data")
for rows in range(1,totalrows+1):
    for cols in range(1,totalcols+1):
        print(sheet.cell(row=rows,column=cols).value,end=" ")
    print()
